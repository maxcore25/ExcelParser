from docxtpl import DocxTemplate
import openpyxl
from docx2pdf import convert
import os
from PyPDF2 import PdfFileMerger

# ===== Part 1 =====
print('===== Part 1 =====')
print()
# Define variable to load the workbook
table_name = input('Enter table name you want to parse: ')
workbook = openpyxl.load_workbook(f"./tables/{table_name}.xlsx")
# Define variable to read the active sheet:
worksheet = workbook.active
# Iterate the loop to read the cell values
for i in range(1, worksheet.max_row):
    vals = []
    for col in worksheet.iter_cols(0, 30):
        vals.append(col[i].value)
    doc = DocxTemplate("Mail_for_700_PP.docx")
    print(vals)
    if 'ИНН' in str(vals[29]):
        spl_vals_29 = vals[29].split(',')
        try:
            fio, inn = spl_vals_29[0], spl_vals_29[1].split(':')[1]
        except IndexError:
            pass
        context = {'fio': fio,
                   'square': vals[11],
                   'inn': inn,
                   'address': vals[13],
                   'cad_num': vals[4],
                   'cad_cost': vals[15]}
    else:
        context = {'fio': vals[29],
                   'square': vals[11],
                   'inn': '-',
                   'address': vals[13],
                   'cad_num': vals[4],
                   'cad_cost': vals[15]}
    doc.render(context)
    doc.save(f"./final_docs/letter_700_PP_{i + 1}.docx")

convert('./final_docs')

# ===== Part 2 =====
print()
print()
print('===== Part 2 =====')
print()
path = os.getcwd() + '\\final_docs\\'
print(path)
all_count = 0
docx_count = 0

for file in os.scandir(path):
    all_count += 1
    filename, file_extension = os.path.splitext(file.name)
    if file_extension == '.docx':
        docx_count += 1
        os.remove(path + file.name)
        print('Removed:', file.name)

print(all_count, docx_count)

# ===== Part 3 =====
print()
print()
print('===== Part 3 =====')
print()
path = os.getcwd() + '\\final_docs\\'
print(path)
counter = 0

merged_files_path = os.getcwd() + f'\\{table_name}\\'
if not os.path.exists(merged_files_path):
    os.makedirs(merged_files_path)

for pdf_letter in os.scandir(path):
    counter += 1
    merger = PdfFileMerger()

    merger.append(path + pdf_letter.name)
    merger.append('Приложение к письму оспаривание КС.pdf')

    # merger.write(os.getcwd() + '\\merged_letters\\' + f'letter_700_PP_{counter + 1}.pdf')
    merger.write(merged_files_path + f'letter_700_PP_{counter + 1}.pdf')
    merger.close()
    print('Merged:', pdf_letter.name)

# ===== Part 4 =====
print()
print()
print('===== Part 4 =====')
print()
path = os.getcwd() + '\\final_docs\\'
print(path)
all_count = 0

for file in os.scandir(path):
    all_count += 1
    os.remove(path + file.name)
    print('Removed:', file.name)

print(all_count)
