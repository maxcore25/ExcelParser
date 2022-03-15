from docxtpl import DocxTemplate
import openpyxl
from docx2pdf import convert

# Define variable to load the workbook
table_name = input('Enter table name you want to parse: ')
workbook = openpyxl.load_workbook(f"./tables/{table_name}.xlsx")
# Define variable to read the active sheet:
worksheet = workbook.active
# Iterate the loop to read the cell values
for i in range(1, worksheet.max_row):
    vals = []
    for col in worksheet.iter_cols(0, 30):
        vals.append(col[i].value)
    doc = DocxTemplate("Mail_for_700_PP.docx")
    print(vals)
    if 'ИНН' in str(vals[29]):
        spl_vals_29 = vals[29].split(',')
        try:
            fio, inn = spl_vals_29[0], spl_vals_29[1].split(':')[1]
        except IndexError:
            pass
        context = {'fio': fio,
                   'square': vals[11],
                   'inn': inn,
                   'address': vals[13],
                   'cad_num': vals[4],
                   'cad_cost': vals[15]}
    else:
        context = {'fio': vals[29],
                   'square': vals[11],
                   'inn': '-',
                   'address': vals[13],
                   'cad_num': vals[4],
                   'cad_cost': vals[15]}
    doc.render(context)
    doc.save(f"./final_docs/letter_700_PP_{i + 1}.docx")

convert('./final_docs')
