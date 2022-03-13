from docxtpl import DocxTemplate
from docx2pdf import convert
import openpyxl
import os
import shutil

from word_killer228 import do_word_nahui
from pdf_merger import do_merge_epta


def do_things(template, table, save_dir):
    os.mkdir(os.path.join(save_dir, 'temp'))
    temp = os.path.join(save_dir, 'temp')
    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(table)
    # Define variable to read the active sheet:
    worksheet = workbook.active
    # Iterate the loop to read the cell values
    for i in range(1, worksheet.max_row):
        vals = []
        for col in worksheet.iter_cols(0, 30):
            vals.append(col[i].value)
        doc = DocxTemplate(template)
        # print(vals)
        if 'ИНН' in str(vals[29]):
            spl_vals_29 = vals[29].split(',')
            try:
                fio, inn = spl_vals_29[0], spl_vals_29[1].split(':')[1]
            except IndexError:
                pass
            context = {'fio': fio,
                       'square': vals[11],
                       'inn': inn,
                       'address': vals[13],
                       'cad_num': vals[4],
                       'cad_cost': vals[15]}
        else:
            context = {'fio': vals[29],
                       'square': vals[11],
                       'inn': '-',
                       'address': vals[13],
                       'cad_num': vals[4],
                       'cad_cost': vals[15]}
        doc.render(context)
        doc.save(os.path.join(temp, f"letter_700_PP_{i + 1}.docx"))

        yield i, worksheet.max_row - 1

    yield 'convert'
    convert(temp)
    do_word_nahui(temp)
    do_merge_epta(temp)
    shutil.make_archive(os.path.join(save_dir, "letters_700_pp"), "zip", temp)
    shutil.rmtree(temp)
