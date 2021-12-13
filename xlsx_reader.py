import openpyxl
# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("120519060Марьина Роща.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
for i in range(1, worksheet.max_row):
    vals = []
    for col in worksheet.iter_cols(1, 6):
        vals.append(col[i].value)
    print(vals[5])

